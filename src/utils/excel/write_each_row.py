import openpyxl
from openpyxl import load_workbook
import os


def write_to_excel(value, filename='output.xlsx'):
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = openpyxl.Workbook()

    sheet = wb.active

    #finde first empty row in cell A
    next_row = sheet.max_row + 1
    while sheet.cell(row=next_row, column=1).value is not None:
        next_row += 1

    sheet.cell(row=next_row, column=1).value = value

    wb.save(filename)